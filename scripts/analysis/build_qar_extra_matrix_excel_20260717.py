#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Build matrix-style Excel sheets for QAR extra experiments.

Input is the long-form CSVs produced by
``scripts/analysis/collect_qar_experiment_results_20260717.py``.

Rows are models, columns are datasets, and each cell stores a compact metric
bundle.  The workbook also keeps long-form sheets and the audit table so
missing/failed jobs remain visible.
"""

from __future__ import annotations

import argparse
import math
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from pandas.errors import EmptyDataError
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATASETS = [
    "dataset5",
    "dataset6",
    "dataset7",
    "dataset8",
    "dataset8-1",
    "dataset9",
    "dataset10",
    "dataset11",
    "dataset12",
    "dataset13",
    "dataset14",
]

FAULT_DESC = {
    "dataset5": "320-感压管路故障",
    "dataset6": "320-HPV活门故障",
    "dataset7": "320-PRV活门故障",
    "dataset8": "320-管道漏气",
    "dataset8-1": "320-PRV活门漏气",
    "dataset9": "321-HPV故障（LEAP）",
    "dataset10": "321-感压管路故障（LEAP）",
    "dataset11": "321-管道漏气",
    "dataset12": "321-PRV故障（LEAP）",
    "dataset13": "787机型空气压缩机故障",
    "dataset14": "777机型PRSOV故障",
}

MODEL_ORDER = [
    "OLinear",
    "xPatch",
    "TimeMixer++",
    "DUET",
    "TimeMixer",
    "TimeXer",
    "Autoformer",
    "Transformer",
    "TimesNet",
    "PatchTST",
    "DLinear",
    "iTransformer",
    "Chronos2",
    "Toto",
    "Moirai",
    "TiRex-2",
    "TiRex",
    "Sundial",
    "TabPFN",
    "MiniROCKET",
    "MultiROCKET",
    "KANAD",
    "AnomalyTransformer",
    "TranAD",
    "USAD",
    "OmniAnomaly",
]


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except EmptyDataError:
        return pd.DataFrame()


def fmt(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ""
        try:
            value = float(value)
        except Exception:
            return value
    try:
        number = float(value)
    except Exception:
        return str(value)
    if math.isinf(number):
        return "inf" if number > 0 else "-inf"
    return f"{number:.4f}"


def clean_dataset_name(value: Any) -> str:
    text = str(value)
    for suffix in [
        "_both_keep50",
        "_both_keep25",
        "_normal_keep50",
        "_normal_keep25",
        "_normalx2",
        "_normalx4",
        "_aug0_1000",
        "_aug0_2000",
        "_aug0_4000",
        "_aug0_19119",
        "_aug0_20000",
    ]:
        if text.endswith(suffix):
            return text[: -len(suffix)]
    return text


def order_models(models: list[str]) -> list[str]:
    order = {name: i for i, name in enumerate(MODEL_ORDER)}
    return sorted(models, key=lambda name: (order.get(name, 999), name))


def nonempty(value: Any) -> bool:
    try:
        if pd.isna(value):
            return False
    except Exception:
        pass
    return str(value).strip() not in {"", "nan", "None"}


def group_label(row: pd.Series, task_name: str) -> str:
    parts = [task_name]
    for key, label in [
        ("variant", "数据规模"),
        ("patch_len", "patch_len"),
        ("history_count", "历史航班数"),
        ("target", "目标变量"),
        ("anchor", "阶段"),
    ]:
        value = row.get(key, "")
        if nonempty(value):
            parts.append(f"{label}={value}")
    return " | ".join(parts)


def group_key_columns(df: pd.DataFrame) -> list[str]:
    keys = ["variant", "patch_len", "history_count", "target", "anchor"]
    return [key for key in keys if key in df.columns]


def latest_success_or_last(sub: pd.DataFrame) -> pd.Series | None:
    if sub.empty:
        return None
    if "status" in sub.columns:
        ok = sub[sub["status"].astype(str).isin(["0", "0.0"])]
        if not ok.empty:
            return ok.iloc[-1]
    return sub.iloc[-1]


def classification_cell(row: pd.Series | None) -> str:
    if row is None:
        return ""
    if str(row.get("status", "")) not in {"0", "0.0", ""}:
        return "FAILED"
    return "\n".join(
        [
            f"acc={fmt(row.get('accuracy', row.get('acc')))}",
            f"macro_f1={fmt(row.get('macro_f1'))}",
            f"weighted_f1={fmt(row.get('weighted_f1'))}",
            f"TN={fmt(row.get('TN'))} FP={fmt(row.get('FP'))}",
            f"FN={fmt(row.get('FN'))} TP={fmt(row.get('TP'))}",
        ]
    )


def forecast_cell(row: pd.Series | None) -> str:
    if row is None:
        return ""
    if str(row.get("status", "")) not in {"0", "0.0", ""}:
        return "FAILED"
    return "\n".join(
        [
            f"mae={fmt(row.get('mae'))}",
            f"mse={fmt(row.get('mse'))}",
            f"rmse={fmt(row.get('rmse'))}",
            f"mape={fmt(row.get('mape'))}",
            f"mspe={fmt(row.get('mspe'))}",
        ]
    )


def anomaly_cell(row: pd.Series | None) -> str:
    if row is None:
        return ""
    if str(row.get("status", "")) not in {"0", "0.0", ""}:
        return "FAILED"
    return "\n".join(
        [
            f"acc={fmt(row.get('accuracy'))}",
            f"f1={fmt(row.get('f1'))}",
            f"precision={fmt(row.get('precision'))}",
            f"recall={fmt(row.get('recall'))}",
            f"TN={fmt(row.get('TN'))} FP={fmt(row.get('FP'))}",
            f"FN={fmt(row.get('FN'))} TP={fmt(row.get('TP'))}",
        ]
    )


def prepare_df(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    out["dataset_base"] = out["dataset"].map(clean_dataset_name)
    for col in ["variant", "patch_len", "history_count", "target", "anchor"]:
        if col not in out.columns:
            out[col] = ""
    return out


def write_dataframe_sheet(wb: Workbook, title: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(title[:31])
    if df.empty:
        ws.cell(1, 1, "EMPTY")
        return
    ws.append(list(df.columns))
    for _, row in df.iterrows():
        ws.append([row.get(c, "") for c in df.columns])
    for col in range(1, len(df.columns) + 1):
        ws.column_dimensions[get_column_letter(col)].width = min(60, max(12, len(str(df.columns[col - 1])) + 2))


def write_matrix_sheet(
    wb: Workbook,
    title: str,
    df: pd.DataFrame,
    task_name: str,
    cell_builder: Callable[[pd.Series | None], str],
) -> None:
    ws = wb.create_sheet(title[:31])
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_group = PatternFill("solid", fgColor="D9EAF7")
    fill_header = PatternFill("solid", fgColor="E2F0D9")
    fill_warn = PatternFill("solid", fgColor="FFF2CC")
    font_title = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
    font_bold = Font(bold=True, name="Microsoft YaHei")
    font_normal = Font(name="Microsoft YaHei", size=10)

    def put(r: int, c: int, value: Any, fill=None, font=None):
        cell = ws.cell(r, c, value)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = font or font_normal
        if fill:
            cell.fill = fill
        return cell

    row_idx = 1
    put(row_idx, 1, title, fill_title, font_title)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DATASETS) + 1)
    row_idx += 2

    put(row_idx, 1, "dataset", fill_header, font_bold)
    for col, dataset in enumerate(DATASETS, 2):
        put(row_idx, col, dataset, fill_header, font_bold)
    row_idx += 1
    put(row_idx, 1, "故障说明", fill_header, font_bold)
    for col, dataset in enumerate(DATASETS, 2):
        put(row_idx, col, FAULT_DESC.get(dataset, ""), fill_header, font_bold)
    row_idx += 2

    if df.empty:
        put(row_idx, 1, "EMPTY")
        return

    df = prepare_df(df)
    key_cols = group_key_columns(df) or ["variant"]

    for _, block in df.groupby(key_cols, dropna=False, sort=False):
        label = group_label(block.iloc[0], task_name)
        put(row_idx, 1, label, fill_group, font_bold)
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DATASETS) + 1)
        row_idx += 1
        put(row_idx, 1, "model", fill_header, font_bold)
        for col, dataset in enumerate(DATASETS, 2):
            put(row_idx, col, dataset, fill_header, font_bold)
        row_idx += 1

        models = order_models(block["model"].dropna().astype(str).unique().tolist())
        for model in models:
            put(row_idx, 1, model, None, font_bold)
            for col, dataset in enumerate(DATASETS, 2):
                sub = block[(block["dataset_base"] == dataset) & (block["model"].astype(str) == model)]
                cell_text = cell_builder(latest_success_or_last(sub))
                fill = fill_warn if cell_text == "FAILED" else None
                put(row_idx, col, cell_text, fill=fill)
            row_idx += 1
        row_idx += 2

    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 28
    for col in range(2, len(DATASETS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 24
    for ridx in range(1, row_idx + 1):
        ws.row_dimensions[ridx].height = 70 if ridx > 4 else 24


def build_workbook(input_dir: Path, output: Path) -> Path:
    cls = read_csv(input_dir / "all_classification_metrics.csv")
    forecast = read_csv(input_dir / "all_forecast_metrics.csv")
    zero = read_csv(input_dir / "all_zero_shot_metrics.csv")
    anomaly = read_csv(input_dir / "all_anomaly_metrics.csv")
    forecast_anomaly = read_csv(input_dir / "all_forecast_anomaly_metrics.csv")
    audit = read_csv(input_dir / "job_audit.csv")
    expected = read_csv(input_dir / "expected_jobs.csv")

    wb = Workbook()
    wb.remove(wb.active)
    write_matrix_sheet(wb, "分类矩阵", cls, "分类", classification_cell)
    write_matrix_sheet(wb, "预测矩阵", forecast, "预测", forecast_cell)
    write_matrix_sheet(wb, "时序大模型矩阵", zero, "时序大模型-单变量预测", forecast_cell)
    write_matrix_sheet(wb, "异常检测矩阵", anomaly, "异常检测", anomaly_cell)
    write_matrix_sheet(wb, "预测头异常检测矩阵", forecast_anomaly, "预测头异常检测", anomaly_cell)
    write_dataframe_sheet(wb, "classification_long", cls)
    write_dataframe_sheet(wb, "forecast_long", forecast)
    write_dataframe_sheet(wb, "zero_shot_long", zero)
    write_dataframe_sheet(wb, "anomaly_long", anomaly)
    write_dataframe_sheet(wb, "forecast_anomaly_long", forecast_anomaly)
    write_dataframe_sheet(wb, "audit", audit)
    write_dataframe_sheet(wb, "expected_jobs", expected)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(build_workbook(args.input_dir, args.output))


if __name__ == "__main__":
    main()
