#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Build a one-sheet QAR Excel containing classification, forecast and anomaly tables."""

from __future__ import annotations

import argparse
import math
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATASETS = [
    "dataset5", "dataset6", "dataset7", "dataset8", "dataset8-1",
    "dataset9", "dataset10", "dataset11", "dataset12", "dataset13", "dataset14",
]
MODELS = ["Transformer", "TimesNet", "PatchTST", "DLinear", "iTransformer"]
FAULT_DESC = {
    "dataset5": "320-感压管路故障",
    "dataset6": "320-HPV活门故障",
    "dataset7": "320-PRV活门故障",
    "dataset8": "320-管道漏气",
    "dataset8-1": "320-PRV活门漏气",
    "dataset9": "321-HPV故障",
    "dataset10": "321-感压管路故障",
    "dataset11": "321-管道漏气",
    "dataset12": "321-PRV故障",
    "dataset13": "787机型空气压缩机故障",
    "dataset14": "777机型PRSOV故障",
}
FORECASTS = [
    ("predict_2_3", "预测结果 2→3（滑行→起飞滑跑，60预测20）"),
    ("predict_4_5", "预测结果 4→5（离地→爬升，60预测20）"),
    ("predict_5_6", "预测结果 5→6（爬升→巡航，60预测20）"),
    ("predict_8_9", "预测结果 8→9（进近→落地，60预测20）"),
]


def load_csv(path: Path) -> pd.DataFrame:
    if path and path.exists():
        return pd.read_csv(path)
    return pd.DataFrame()


def clean(value):
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        if math.isinf(value):
            return "inf" if value > 0 else "-inf"
        return round(value, 6)
    return value


def order_models(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty or "model" not in df.columns:
        return df
    order = {m: i for i, m in enumerate(MODELS)}
    out = df.copy()
    out["_model_order"] = out["model"].map(order).fillna(999)
    sort_cols = [c for c in ["dataset", "_model_order"] if c in out.columns]
    out = out.sort_values(sort_cols).drop(columns=["_model_order"])
    return out


def build_excel(
    base_dir: Path,
    anomaly_dir: Path | None,
    split_manifest: Path | None,
    output: Path,
    title_suffix: str,
):
    cls = order_models(load_csv(base_dir / "classification" / "all_metrics.csv"))
    forecasts = {
        name: order_models(load_csv(base_dir / f"forecast_{name}" / "all_metrics.csv"))
        for name, _ in FORECASTS
    }
    anomaly = order_models(load_csv(anomaly_dir / "all_anomaly_metrics.csv")) if anomaly_dir else pd.DataFrame()
    split = load_csv(split_manifest) if split_manifest else load_csv(base_dir / "manifests" / "split_time_ranges.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "results"

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_sub = PatternFill("solid", fgColor="D9EAF7")
    fill_header = PatternFill("solid", fgColor="E2F0D9")
    fill_warn = PatternFill("solid", fgColor="FFF2CC")
    white = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
    bold = Font(bold=True, name="Microsoft YaHei")
    normal = Font(name="Microsoft YaHei")

    def write_row(row_idx: int, values, fill=None, font=None):
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=clean(value))
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            cell.font = font or normal
            if fill:
                cell.fill = fill

    row = 1
    write_row(row, [f"实验条件{title_suffix}"], fill_title, white)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    row += 1
    conditions = [
        ["数据来源", "tsfile compact cache；分类与预测均读取 qar_compact_shiftN80.npz"],
        ["分类工况", "dataset5~12 使用 320/321 多锚点工况，删除 6→8 巡航→进近；dataset13/14 使用各自适配字段"],
        ["预测工况", "按 2→3、4→5、5→6、8→9 四个阶段转换分别建预测小数据集；60 个点预测后 20 个点"],
        ["分类训练", "CrossEntropyLoss 支持 balanced class weight；early stopping 默认 macro_f1"],
        ["异常检测", "one-class reconstruction；训练/验证只用正常类 0；p95 阈值来自正常验证集，不用测试集定阈值"],
        ["模型", "、".join(MODELS)],
    ]
    if not split.empty and "task" in split.columns:
        strategy_hint = "per-class chronological 7:1:2" if (split["task"].astype(str).str.contains("forecast").any()) else "split manifest"
        conditions.append(["切分记录", f"{strategy_hint}；详见 split_time_ranges.csv"])
    for item in conditions:
        write_row(row, item)
        row += 1
    row += 1

    cls_cols = ["model", "acc", "accuracy", "macro_f1", "weighted_f1", "true_counts", "pred_counts", "TN", "FP", "FN", "TP"]
    forecast_cols = ["model", "mae", "mse", "rmse", "mape", "mspe"]
    anomaly_cols = ["model", "accuracy", "precision", "recall", "f1", "true_counts", "pred_counts", "TN", "FP", "FN", "TP"]

    for dataset in DATASETS:
        write_row(row, ["数据集说明"], fill_title, white)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        row += 1
        write_row(row, ["dataset", dataset])
        row += 1
        write_row(row, ["故障说明", FAULT_DESC.get(dataset, "")])
        row += 1

        if not split.empty:
            sub_split = split[(split["task"] == "classification") & (split["dataset"] == dataset)]
            for _, sr in sub_split.iterrows():
                fill = fill_warn if sr.get("split") == "test" and int(sr.get("class1", 0)) <= 0 else None
                write_row(row, [
                    f"{sr['split']} 样本", int(sr["n"]),
                    "class0", int(sr["class0"]),
                    "class1", int(sr["class1"]),
                    "时间范围", f"{sr['first_time_key']} ~ {sr['last_time_key']}",
                ], fill=fill)
                row += 1
        row += 1

        write_row(row, ["分类结果"], fill_sub, bold)
        row += 1
        write_row(row, cls_cols, fill_header, bold)
        row += 1
        sub = cls[cls["dataset"] == dataset] if not cls.empty else pd.DataFrame()
        for _, metric in sub.iterrows():
            write_row(row, [metric.get(c, "") for c in cls_cols])
            row += 1
        row += 1

        for forecast_name, forecast_title in FORECASTS:
            write_row(row, [forecast_title], fill_sub, bold)
            row += 1
            write_row(row, forecast_cols, fill_header, bold)
            row += 1
            fdf = forecasts[forecast_name]
            sub_f = fdf[fdf["dataset"] == dataset] if not fdf.empty else pd.DataFrame()
            for _, metric in sub_f.iterrows():
                write_row(row, [metric.get(c, "") for c in forecast_cols])
                row += 1
            row += 1

        if not anomaly.empty:
            write_row(row, ["异常检测结果 p95（one-class，训练不用故障样本）"], fill_sub, bold)
            row += 1
            write_row(row, anomaly_cols, fill_header, bold)
            row += 1
            sub_a = anomaly[anomaly["dataset"] == dataset]
            for _, metric in sub_a.iterrows():
                write_row(row, [metric.get(c, "") for c in anomaly_cols])
                row += 1
            row += 1

        row += 2

    for col in range(1, 13):
        width = 14
        if col == 1:
            width = 22
        elif col in (2, 6, 7, 8):
            width = 24
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base-dir", type=Path, required=True)
    parser.add_argument("--anomaly-dir", type=Path)
    parser.add_argument("--split-manifest", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--title-suffix", default="")
    args = parser.parse_args()
    print(build_excel(args.base_dir, args.anomaly_dir, args.split_manifest, args.output, args.title_suffix))


if __name__ == "__main__":
    main()
