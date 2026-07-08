"""Build a one-sheet Excel report for QAR custom-condition experiments.

This script merges:

- classification metrics with true/pred counts and TN/FP/FN/TP
- phase-start-80 forecasting metrics

into a single worksheet.  It intentionally keeps code in the repository root
area and writes only the generated report into ``experiment_artifacts``.
"""

from __future__ import annotations

import argparse
import math
import os
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


DATASET_NAMES = {
    "dataset5": "320-感压管路故障",
    "dataset6": "320-HPV活门故障",
    "dataset7": "320-PRV活门故障",
    "dataset8": "320-管道漏气",
    "dataset8-1": "320-PRV活门漏气",
    "dataset9": "321-HPV故障",
    "dataset10": "321-感压管路故障",
    "dataset11": "321-管道漏气",
    "dataset12": "321-PRV故障",
    "dataset12_aug0": "321-PRV故障（0类追加CSV数据版）",
    "dataset13": "787机型空气压缩机故障",
    "dataset14": "777机型PRSOV故障",
}

DATASET_ORDER = [
    "dataset5",
    "dataset6",
    "dataset7",
    "dataset8",
    "dataset8-1",
    "dataset9",
    "dataset10",
    "dataset11",
    "dataset12",
    "dataset12_aug0",
    "dataset13",
    "dataset14",
]

MODEL_ORDER = ["Transformer", "TimesNet", "PatchTST", "DLinear", "iTransformer"]


TITLE_FONT = Font(bold=True, size=12)
HEADER_FONT = Font(bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="E2F0D9")


def clean_value(value: Any) -> Any:
    """Keep numbers readable in Excel while preserving inf/count strings."""
    if pd.isna(value):
        return ""
    if isinstance(value, float):
        if math.isinf(value):
            return "inf" if value > 0 else "-inf"
        return round(value, 6)
    return value


def load_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        raise FileNotFoundError(path)
    return pd.read_csv(path)


def order_frame(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_dataset_order"] = df["dataset"].map({name: i for i, name in enumerate(DATASET_ORDER)})
    df["_model_order"] = df["model"].map({name: i for i, name in enumerate(MODEL_ORDER)})
    return df.sort_values(["_dataset_order", "_model_order"]).drop(
        columns=["_dataset_order", "_model_order"], errors="ignore"
    )


def write_row(ws, row_idx: int, values: list[Any], *, font=None, fill=None) -> None:
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row_idx, col_idx, clean_value(value))
        if font is not None:
            cell.font = font
        if fill is not None:
            cell.fill = fill
        cell.alignment = Alignment(vertical="center")


def write_section_title(ws, row_idx: int, title: str, width: int) -> None:
    write_row(ws, row_idx, [title], font=TITLE_FONT, fill=SECTION_FILL)
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=width)


def write_table(ws, row_idx: int, header: list[str], rows: list[list[Any]]) -> int:
    write_row(ws, row_idx, header, font=HEADER_FONT, fill=HEADER_FILL)
    row_idx += 1
    for row in rows:
        write_row(ws, row_idx, row)
        row_idx += 1
    return row_idx


def build_report(
    classification_csv: Path,
    forecast_phase80_csv: Path,
    output_xlsx: Path,
) -> Path:
    cls = order_frame(load_csv(classification_csv))
    phase80 = order_frame(load_csv(forecast_phase80_csv))

    cls = cls[cls["dataset"].isin(DATASET_ORDER)]
    phase80 = phase80[phase80["dataset"].isin(DATASET_ORDER)]

    phase80_metrics = {
        (row["dataset"], row["model"]): row
        for _, row in phase80.iterrows()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "results"
    ws.freeze_panes = "A1"

    row_idx = 1
    info_width = 15
    write_section_title(ws, row_idx, "实验条件", info_width)
    row_idx += 1
    info_rows = [
        ["分类数据", "新工况 compact cache；除 dataset13 外参考 320321gongkuang.py，dataset13 参考 build_dataset15_1.py。"],
        ["分类训练", "class_weight=balanced；early_stop_metric=macro_f1；二分类正类=1类故障/异常。"],
        ["分类结果备注", "当前分类代码用 TEST 同时作为 validation/test 做 early stopping，分类指标会偏乐观；严格结果应重新划分 TRAIN/VAL/TEST 后再跑。"],
        ["预测 phase80 工况", "在 0→12 飞行阶段开始位置取 80 个点，前 60 预测后 20。"],
        ["预测说明", "mape/mspe 出现 inf 是因为原始/归一化目标中存在 0，主要看 mae/mse/rmse。"],
        ["混淆矩阵定义", "TN=真实0预测0；FP=真实0预测1；FN=真实1预测0；TP=真实1预测1。"],
    ]
    row_idx = write_table(ws, row_idx, ["item", "value"], info_rows)
    row_idx += 2

    cls_header = [
        "model",
        "acc",
        "accuracy",
        "macro_f1",
        "weighted_f1",
        "true_counts",
        "pred_counts",
        "TN",
        "FP",
        "FN",
        "TP",
    ]
    forecast_header = [
        "model",
        "mae",
        "mse",
        "rmse",
        "mape",
        "mspe",
    ]

    for dataset in DATASET_ORDER:
        dataset_cls = cls[cls["dataset"] == dataset]
        if dataset_cls.empty:
            continue

        write_section_title(ws, row_idx, "数据集说明", info_width)
        row_idx += 1
        row_idx = write_table(
            ws,
            row_idx,
            ["dataset", dataset],
            [["故障说明", DATASET_NAMES.get(dataset, dataset)]],
        )
        row_idx += 1

        write_section_title(ws, row_idx, "分类结果", info_width)
        row_idx += 1
        cls_rows = []
        for _, r in dataset_cls.iterrows():
            cls_rows.append(
                [
                    r["model"],
                    r["acc"],
                    r["accuracy"],
                    r["macro_f1"],
                    r["weighted_f1"],
                    r.get("true_counts", ""),
                    r.get("pred_counts", ""),
                    r.get("TN", ""),
                    r.get("FP", ""),
                    r.get("FN", ""),
                    r.get("TP", ""),
                ]
            )
        row_idx = write_table(ws, row_idx, cls_header, cls_rows)
        row_idx += 2

        write_section_title(ws, row_idx, "预测结果", info_width)
        row_idx += 1
        forecast_rows = []
        for model in MODEL_ORDER:
            p = phase80_metrics.get((dataset, model))
            forecast_rows.append(
                [
                    model,
                    "" if p is None else p.get("mae", ""),
                    "" if p is None else p.get("mse", ""),
                    "" if p is None else p.get("rmse", ""),
                    "" if p is None else p.get("mape", ""),
                    "" if p is None else p.get("mspe", ""),
                ]
            )
        row_idx = write_table(ws, row_idx, forecast_header, forecast_rows)
        row_idx += 3

    for col_idx in range(1, info_width + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 16
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 18

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(output_xlsx.parent)) as tmp:
        tmp_path = Path(tmp.name)
    final_path = output_xlsx
    try:
        wb.save(tmp_path)
        try:
            os.replace(tmp_path, output_xlsx)
        except PermissionError:
            final_path = output_xlsx.with_name(f"{output_xlsx.stem}_one_sheet{output_xlsx.suffix}")
            os.replace(tmp_path, final_path)
    finally:
        if tmp_path.exists():
            tmp_path.unlink()
    if final_path != output_xlsx:
        print(f"Target file is locked; wrote fallback report to: {final_path}")
    return final_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--classification-csv",
        type=Path,
        default=Path("experiment_artifacts/QAR_custom_conditions_classification_20260708/all_metrics.csv"),
    )
    parser.add_argument(
        "--forecast-phase80-csv",
        type=Path,
        default=Path("experiment_artifacts/QAR_custom_conditions_forecast_phase80_20260708/all_metrics.csv"),
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path("experiment_artifacts/QAR_custom_conditions_summary_20260708/QAR_custom_conditions_results_20260708_phase80_only.xlsx"),
    )
    args = parser.parse_args()
    final_path = build_report(
        args.classification_csv,
        args.forecast_phase80_csv,
        args.output_xlsx,
    )
    print(final_path)


if __name__ == "__main__":
    main()
