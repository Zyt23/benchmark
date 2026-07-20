#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Build the final QAR benchmark as one UTF-8-safe Excel worksheet.

The input directory contains long-form CSV files named:

* all_classification_metrics.csv
* all_forecast_metrics.csv
* all_zero_shot_metrics.csv
* all_anomaly_metrics.csv
* split_audit.csv (optional)

Rows are models, columns are datasets, and each result cell is a compact metric
bundle.  Dataset-12 augmentation variants and the four forecast anchors are
written as separate blocks on the same worksheet.
"""

from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Any, Callable

import pandas as pd
from pandas.errors import EmptyDataError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATASETS = [
    "dataset5", "dataset6", "dataset7", "dataset8", "dataset8-1",
    "dataset9", "dataset10", "dataset11", "dataset12", "dataset13",
    "dataset14",
]

FAULT_DESC = {
    "dataset5": "320-感压管路故障",
    "dataset6": "320-HPV活门故障",
    "dataset7": "320-PRV活门故障",
    "dataset8": "320-管道漏气",
    "dataset8-1": "320-PRV活门漏气",
    "dataset9": "321-HPV故障",
    "dataset10": "321-感压管路故障",
    "dataset11": "321-管道漏气",
    "dataset12": "321-PRV故障",
    "dataset13": "787机型空气压缩机故障",
    "dataset14": "777机型PRSOV故障",
}

MODEL_ORDER = [
    "OLinear", "xPatch", "TimeMixer++", "DUET", "TimeMixer", "TimeXer",
    "iTransformer", "DLinear", "PatchTST", "TimesNet", "Transformer",
    "Autoformer", "TiRex-2", "TiRex", "Chronos-2", "Toto-2.0", "Moirai", "Sundial",
    "MambaSL", "VSFormer", "LITE", "MultiROCKET", "MiniROCKET",
    "TabPFN", "KANAD", "AnomalyTransformer", "TranAD", "USAD",
    "OmniAnomaly",
]

FULL_SHOT_MODELS = [
    "OLinear", "xPatch", "TimeMixer++", "DUET", "TimeMixer", "TimeXer",
    "iTransformer", "DLinear", "PatchTST", "TimesNet", "Autoformer",
    "Transformer",
]

ZERO_SHOT_MODELS = ["TiRex-2", "TiRex", "Chronos-2", "Toto-2.0", "Moirai", "Sundial"]

MODEL_DISPLAY = {
    "Chronos2": "Chronos-2",
    "Toto": "Toto-2.0",
    "TiRex2": "TiRex-2",
}

UNAVAILABLE_MODELS = {
    "TimeMixer++": "UNAVAILABLE\n官方仓库尚未发布 TimeMixer++ 实现",
    "TiRex-2": "GATED_WEIGHT\n需接受 Hugging Face 条款并使用独立 torch>=2.8 环境",
}

AUGMENT_SUFFIXES = [
    "_aug0_2000", "_aug0_4000", "_aug0_6000", "_aug0_10000",
    "_aug0_19119", "_aug0_20000", "_aug0_1000", "_normal_keep25",
    "_normal_keep50", "_both_keep25", "_both_keep50", "_normalx2",
    "_normalx4",
]

MOJIBAKE_MARKERS = (
    "锟斤拷", "鎰熷帇", "鏁呴殰", "棰勬祴", "鍒嗙被", "绫诲埆",
    "妫€娴", "璁粌", "楠岃瘉", "娴嬭瘯", "鏁版嵁闆",
)


def read_csv(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        return pd.read_csv(path)
    except EmptyDataError:
        return pd.DataFrame()


def fmt(value: Any, digits: int = 4) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ""
        try:
            value = float(value)
        except ValueError:
            return value
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if math.isinf(number):
        return "inf" if number > 0 else "-inf"
    if number.is_integer() and abs(number) >= 1:
        return str(int(number))
    return f"{number:.{digits}f}"


def base_dataset(value: Any) -> str:
    text = str(value).strip()
    for suffix in AUGMENT_SUFFIXES:
        if text.endswith(suffix):
            return text[: -len(suffix)]
    return text


def infer_variant(row: pd.Series) -> str:
    dataset = str(row.get("dataset", ""))
    for suffix in AUGMENT_SUFFIXES:
        if dataset.endswith(suffix):
            return suffix.removeprefix("_")
    variant = str(row.get("variant", "")).strip()
    if variant and variant.lower() != "nan":
        return variant
    return "base"


def infer_anchor(row: pd.Series) -> str:
    anchor = str(row.get("anchor", "")).strip()
    if anchor and anchor.lower() != "nan":
        return anchor.replace("predict_", "")
    run_tag = str(row.get("run_tag", ""))
    match = re.search(r"predict_(\d+_\d+)", run_tag)
    return match.group(1) if match else ""


def normalize(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    out = df.copy()
    if "dataset" not in out.columns:
        return pd.DataFrame()
    out["dataset_base"] = out["dataset"].map(base_dataset)
    out["variant_key"] = out.apply(infer_variant, axis=1)
    out["anchor_key"] = out.apply(infer_anchor, axis=1)
    if "model" not in out.columns:
        out["model"] = ""
    out["model"] = out["model"].replace(MODEL_DISPLAY)
    return out


def model_order(values: list[str]) -> list[str]:
    rank = {name: idx for idx, name in enumerate(MODEL_ORDER)}
    return sorted(values, key=lambda item: (rank.get(item, 999), item))


def successful(row: pd.Series) -> bool:
    value = str(row.get("status", "")).strip().lower()
    return value in {"", "0", "0.0", "ok", "success", "nan"}


def latest(subset: pd.DataFrame) -> pd.Series | None:
    if subset.empty:
        return None
    good = subset[subset.apply(successful, axis=1)]
    return (good if not good.empty else subset).iloc[-1]


def classification_cell(row: pd.Series | None) -> str:
    if row is None:
        return ""
    if not successful(row):
        return "FAILED"
    acc = row.get("acc", row.get("accuracy"))
    return (
        f"acc={fmt(acc)}  macro_f1={fmt(row.get('macro_f1'))}\n"
        f"weighted_f1={fmt(row.get('weighted_f1'))}\n"
        f"true={fmt(row.get('true_counts'))}  pred={fmt(row.get('pred_counts'))}\n"
        f"TN={fmt(row.get('TN'))} FP={fmt(row.get('FP'))} "
        f"FN={fmt(row.get('FN'))} TP={fmt(row.get('TP'))}"
    )


def forecast_cell(row: pd.Series | None) -> str:
    if row is None:
        return ""
    if not successful(row):
        return "FAILED"
    return (
        f"MAE={fmt(row.get('mae'))}  MSE={fmt(row.get('mse'))}\n"
        f"RMSE={fmt(row.get('rmse'))}\n"
        f"MAPE={fmt(row.get('mape'))}  MSPE={fmt(row.get('mspe'))}"
    )


def anomaly_cell(row: pd.Series | None) -> str:
    if row is None:
        return ""
    if not successful(row):
        return "FAILED"
    return (
        f"acc={fmt(row.get('accuracy'))}  bal_acc={fmt(row.get('balanced_accuracy'))}\n"
        f"F1={fmt(row.get('f1'))}  macro_f1={fmt(row.get('macro_f1'))}\n"
        f"P={fmt(row.get('precision'))} R={fmt(row.get('recall'))} "
        f"ROC={fmt(row.get('roc_auc', row.get('auroc')))} "
        f"PR={fmt(row.get('pr_auc', row.get('auprc')))}\n"
        f"TN={fmt(row.get('TN'))} FP={fmt(row.get('FP'))} "
        f"FN={fmt(row.get('FN'))} TP={fmt(row.get('TP'))}"
    )


def build_workbook(input_dir: Path, output: Path) -> Path:
    classification = normalize(read_csv(input_dir / "all_classification_metrics.csv"))
    forecast = normalize(read_csv(input_dir / "all_forecast_metrics.csv"))
    zero_shot = normalize(read_csv(input_dir / "all_zero_shot_metrics.csv"))
    anomaly = normalize(read_csv(input_dir / "all_anomaly_metrics.csv"))
    forecast_anomaly = normalize(read_csv(input_dir / "all_forecast_anomaly_metrics.csv"))
    coverage_audit = read_csv(input_dir / "coverage_audit.csv")
    anomaly_correlation = read_csv(input_dir / "forecast_anomaly_correlations.csv")
    split_audit = read_csv(input_dir / "split_audit.csv")
    shortcut_audit = read_csv(input_dir / "shortcut_audit_summary.csv")
    augmentation_manifest = read_csv(input_dir / "dataset12_augmentation_manifest.csv")

    wb = Workbook()
    ws = wb.active
    ws.title = "全部结果"

    thin = Side(style="thin", color="B4C6E7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_fill = PatternFill("solid", fgColor="1F4E78")
    group_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    title_font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
    bold_font = Font(name="Microsoft YaHei", bold=True)
    normal_font = Font(name="Microsoft YaHei", size=10)

    def put(row: int, col: int, value: Any, *, fill=None, font=None) -> None:
        cell = ws.cell(row, col, value)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.font = font or normal_font
        if fill is not None:
            cell.fill = fill

    row = 1
    put(row, 1, "QAR 全部实验结果（无测试集选模版）", fill=title_fill, font=title_font)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
    row += 2
    conditions = [
        "划分：每个类别内部按航班时间排序后 7:1:2；训练/验证/测试航班源文件不重叠；测试集同时含正常与故障。",
        "分类：使用新工况长序列拼接，已移除 6→8 锚点；balanced CrossEntropy；按验证集 macro_f1 早停，测试集仅在最终评估一次。",
        "预测：四个工况分别建集（2→3、4→5、5→6、8→9），转换点前 30 点+后 50 点；前 60 点预测后 20 点。",
        "预测头异常检测：只用正常训练集拟合；误差类型和阈值只在混合验证集上选择；测试集仅最终评估一次。",
    ]
    put(row, 1, "实验条件", fill=group_fill, font=bold_font)
    put(row, 2, "\n".join(conditions), fill=group_fill)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=len(DATASETS) + 1)
    ws.row_dimensions[row].height = 88
    row += 2

    put(row, 1, "dataset", fill=header_fill, font=bold_font)
    for column, dataset in enumerate(DATASETS, 2):
        put(row, column, dataset, fill=header_fill, font=bold_font)
    row += 1
    put(row, 1, "故障说明", fill=header_fill, font=bold_font)
    for column, dataset in enumerate(DATASETS, 2):
        put(row, column, FAULT_DESC[dataset], fill=header_fill, font=bold_font)
    row += 2

    def write_block(title: str, frame: pd.DataFrame,
                    builder: Callable[[pd.Series | None], str],
                    group_col: str | None = None,
                    expected_models: list[str] | None = None) -> None:
        nonlocal row
        put(row, 1, title, fill=title_fill, font=title_font)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
        row += 1
        if frame.empty:
            put(row, 1, "尚未运行")
            row += 2
            return

        groups: list[tuple[str, pd.DataFrame]]
        if group_col is None:
            groups = [("", frame)]
        else:
            groups = [(str(key), block) for key, block in frame.groupby(group_col, sort=False, dropna=False)]

        for label, block in groups:
            if label:
                display = label.replace("_", "→") if group_col == "anchor_key" else label
                put(row, 1, f"{title}｜{display}", fill=group_fill, font=bold_font)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
                row += 1
            put(row, 1, "model", fill=header_fill, font=bold_font)
            for column, dataset in enumerate(DATASETS, 2):
                put(row, column, dataset, fill=header_fill, font=bold_font)
            row += 1
            present_models = block["model"].dropna().astype(str).unique().tolist()
            models = model_order(list(set(present_models) | set(expected_models or [])))
            for model in models:
                put(row, 1, model, font=bold_font)
                for column, dataset in enumerate(DATASETS, 2):
                    subset = block[(block["dataset_base"] == dataset) & (block["model"].astype(str) == model)]
                    selected = latest(subset)
                    text = UNAVAILABLE_MODELS.get(model, "") if selected is None else builder(selected)
                    put(row, column, text, fill=warn_fill if text == "FAILED" else None)
                ws.row_dimensions[row].height = 78
                row += 1
            row += 2

    group_text = classification.get("experiment_group", pd.Series("", index=classification.index)).fillna("").astype(str)
    base_cls = classification[(classification["variant_key"] == "base") & ~group_text.str.contains("patchlen")] if not classification.empty else classification
    aug_cls = classification[classification["variant_key"].str.startswith("aug0_")] if not classification.empty else classification
    scale_cls = classification[classification["variant_key"].isin(
        ["both_keep25", "both_keep50", "normal_keep25", "normal_keep50"])] if not classification.empty else classification
    normal_aug_cls = classification[classification["variant_key"].isin(
        ["normalx2", "normalx4"])] if not classification.empty else classification
    patch_cls = classification[group_text.str.contains("patchlen")] if not classification.empty else classification
    write_block("故障分类", base_cls, classification_cell)
    write_block("dataset12 追加正常样本分类", aug_cls, classification_cell, "variant_key")
    write_block("分类：数据规模缩减", scale_cls, classification_cell, "variant_key")
    write_block("分类：正常样本扩增", normal_aug_cls, classification_cell, "variant_key")
    write_block("分类：PatchTST patch 长度", patch_cls, classification_cell, "patch_len")
    if not augmentation_manifest.empty:
        put(row, 1, "dataset12 追加正常样本清单", fill=title_fill, font=title_font)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
        row += 1
        put(row, 1, "variant", fill=header_fill, font=bold_font)
        for column, dataset in enumerate(DATASETS, 2):
            put(row, column, dataset, fill=header_fill, font=bold_font)
        row += 1
        for _, manifest_row in augmentation_manifest.sort_values("requested_added_normal").iterrows():
            put(row, 1, str(manifest_row.get("variant", "")), font=bold_font)
            text = (
                f"请求追加={fmt(manifest_row.get('requested_added_normal'))}\n"
                f"有效追加={fmt(manifest_row.get('effective_added_normal'))}  "
                f"未转换={fmt(manifest_row.get('rejected_or_missing'))}\n"
                f"总样本={fmt(manifest_row.get('total_samples'))}  "
                f"class0={fmt(manifest_row.get('class0_count'))}  "
                f"class1={fmt(manifest_row.get('class1_count'))}\n"
                f"源文件去重数={fmt(manifest_row.get('unique_sources'))}"
            )
            put(row, DATASETS.index("dataset12") + 2, text)
            ws.row_dimensions[row].height = 70
            row += 1
        row += 2
    forecast_group = forecast.get("experiment_group", pd.Series("", index=forecast.index)).fillna("").astype(str)
    base_forecast = forecast[(forecast["variant_key"] == "base") & ~forecast_group.str.contains("patchlen")]
    scale_forecast = forecast[forecast["variant_key"].isin(["both_keep25", "both_keep50"])].copy()
    normal_aug_forecast = forecast[forecast["variant_key"].isin(["normalx2", "normalx4"])].copy()
    for frame in (scale_forecast, normal_aug_forecast):
        if not frame.empty:
            frame["variant_anchor"] = frame["variant_key"].astype(str) + " / " + frame["anchor_key"].astype(str)
    patch_forecast = forecast[forecast_group.str.contains("patchlen")].copy()
    if not patch_forecast.empty:
        patch_forecast["anchor_patch"] = patch_forecast["anchor_key"].astype(str) + " / patch=" + patch_forecast["patch_len"].astype(str)
    write_block("预测性维护（全监督）", base_forecast, forecast_cell, "anchor_key", FULL_SHOT_MODELS)
    write_block("预测：数据规模缩减", scale_forecast, forecast_cell, "variant_anchor")
    write_block("预测：正常样本扩增", normal_aug_forecast, forecast_cell, "variant_anchor")
    write_block("预测：PatchTST/TimeXer patch 长度", patch_forecast, forecast_cell, "anchor_patch", ["PatchTST", "TimeXer"])

    zero_group = zero_shot.get("experiment_group", pd.Series("", index=zero_shot.index)).fillna("").astype(str)
    zero_base = zero_shot[~zero_group.str.contains("foundation_context40|univariate_foundation", regex=True)]
    foundation_context = zero_shot[zero_group.str.contains("foundation_context40")].copy()
    univariate = zero_shot[zero_group.str.contains("univariate_foundation")].copy()
    if not foundation_context.empty:
        foundation_context["history_key"] = "history=" + foundation_context["history_count"].astype(str) + " flights"
    write_block(
        "预测性维护（零样本时序大模型）",
        zero_base,
        forecast_cell,
        "anchor_key",
        ZERO_SHOT_MODELS,
    )
    write_block("大模型：历史航班上下文", foundation_context, forecast_cell, "history_key", ["TiRex-2", "Chronos-2", "Toto-2.0", "Moirai"])
    write_block("单变量大模型：总管压力", univariate, forecast_cell, "anchor_key", ["Sundial"])
    write_block("时序异常检测（纯单类 P95）", anomaly, anomaly_cell)
    write_block("预测模型更换异常检测头", forecast_anomaly, anomaly_cell, "anchor_key", ["Transformer", "TimesNet", "PatchTST", "DLinear", "iTransformer"])

    def write_plain_table(title: str, frame: pd.DataFrame) -> None:
        nonlocal row
        if frame.empty:
            return
        put(row, 1, title, fill=title_fill, font=title_font)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max(len(DATASETS) + 1, len(frame.columns)))
        row += 1
        for column, name in enumerate(frame.columns, 1):
            put(row, column, str(name), fill=header_fill, font=bold_font)
        row += 1
        for _, table_row in frame.iterrows():
            for column, value in enumerate(table_row.tolist(), 1):
                put(row, column, fmt(value))
            row += 1
        row += 2

    write_plain_table("实验覆盖审计", coverage_audit)
    write_plain_table("预测误差与异常检测性能相关性（forecast_quality=-MSE）", anomaly_correlation)

    if not shortcut_audit.empty:
        put(row, 1, "传感器可分性与元信息捷径审计", fill=title_fill, font=title_font)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
        row += 1
        put(row, 1, "diagnostic", fill=header_fill, font=bold_font)
        for column, dataset in enumerate(DATASETS, 2):
            put(row, column, dataset, fill=header_fill, font=bold_font)
        row += 1
        for diagnostic in ["signal-only 固定 LR", "metadata-only 固定 LR", "最强单传感器效应量"]:
            put(row, 1, diagnostic, font=bold_font)
            for column, dataset in enumerate(DATASETS, 2):
                subset = shortcut_audit[shortcut_audit["dataset"].astype(str) == dataset]
                if subset.empty:
                    text = ""
                else:
                    audit = subset.iloc[-1]
                    if diagnostic.startswith("signal"):
                        text = (
                            f"acc={fmt(audit.get('signal_accuracy'))} "
                            f"macro_f1={fmt(audit.get('signal_macro_f1'))}\n"
                            f"ROC={fmt(audit.get('signal_roc_auc'))}\n"
                            f"TN={fmt(audit.get('signal_TN'))} FP={fmt(audit.get('signal_FP'))} "
                            f"FN={fmt(audit.get('signal_FN'))} TP={fmt(audit.get('signal_TP'))}"
                        )
                    elif diagnostic.startswith("metadata"):
                        text = (
                            f"acc={fmt(audit.get('metadata_accuracy'))} "
                            f"macro_f1={fmt(audit.get('metadata_macro_f1'))}\n"
                            f"ROC={fmt(audit.get('metadata_roc_auc'))}\n"
                            f"TN={fmt(audit.get('metadata_TN'))} FP={fmt(audit.get('metadata_FP'))} "
                            f"FN={fmt(audit.get('metadata_FN'))} TP={fmt(audit.get('metadata_TP'))}"
                        )
                    else:
                        text = (
                            f"{audit.get('top_sensor', '')}\n"
                            f"|d| train={fmt(audit.get('top_sensor_train_abs_d'))} "
                            f"test={fmt(audit.get('top_sensor_test_abs_d'))}\n"
                            f"方向一致={audit.get('top_sensor_direction_consistent', '')}"
                        )
                put(row, column, text)
            ws.row_dimensions[row].height = 70
            row += 1
        row += 2

    if not split_audit.empty:
        put(row, 1, "数据划分审计", fill=title_fill, font=title_font)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
        row += 1
        for column, name in enumerate(split_audit.columns, 1):
            put(row, column, name, fill=header_fill, font=bold_font)
        row += 1
        for _, audit_row in split_audit.iterrows():
            for column, value in enumerate(audit_row.tolist(), 1):
                put(row, column, fmt(value))
            row += 1

    ws.freeze_panes = "B7"
    ws.column_dimensions["A"].width = 30
    for column in range(2, len(DATASETS) + 2):
        ws.column_dimensions[get_column_letter(column)].width = 31
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    # Reopen and fail fast if replacement characters or question-mark headings
    # slipped into the workbook through a bad terminal encoding conversion.
    check = load_workbook(output, read_only=True, data_only=True)
    bad: list[str] = []
    for sheet in check.worksheets:
        for values in sheet.iter_rows(values_only=True):
            for value in values:
                if isinstance(value, str):
                    stripped = value.strip()
                    if (
                        "�" in value
                        or re.fullmatch(r"\?{2,}", stripped)
                        or any(marker in value for marker in MOJIBAKE_MARKERS)
                    ):
                        bad.append(value)
    check.close()
    if bad:
        raise RuntimeError(f"workbook contains mojibake markers: {bad[:5]}")
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    print(build_workbook(args.input_dir, args.output))


if __name__ == "__main__":
    main()
