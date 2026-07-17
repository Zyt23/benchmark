#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Build a one-sheet benchmark matrix for QAR experiments.

The report layout follows the requested table:

* columns are datasets;
* rows are models;
* four blocks are written on one sheet:
  Predictive Maintenance (Full-shot), Predictive Maintenance (Zero-shot),
  Fault classification, and Anomaly detection.

Each model/dataset cell stores a compact metric bundle.  Missing results are
kept explicit with a reason such as ``NOT_IMPLEMENTED`` or ``MISSING_DEP``.
"""

from __future__ import annotations

import argparse
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DATASETS = [
    "dataset5",
    "dataset6",
    "dataset7",
    "dataset8",
    "dataset8-1",
    "dataset9",
    "dataset10",
    "dataset11",
    "dataset12",
    "dataset13",
    "dataset14",
]

FAULT_DESC = {
    "dataset5": "320-感压管路故障",
    "dataset6": "320-HPV活门故障",
    "dataset7": "320-PRV活门故障",
    "dataset8": "320-管道漏气",
    "dataset8-1": "320-PRV活门漏气",
    "dataset9": "321-HPV故障",
    "dataset10": "321-感压管路故障",
    "dataset11": "321-管道漏气",
    "dataset12": "321-PRV故障",
    "dataset13": "787机型空气压缩机故障",
    "dataset14": "777机型PRSOV故障",
}

FULL_SHOT_MODELS = [
    "OLinear",
    "xPatch",
    "TimeMixer++",
    "DUET",
    "TimeMixer",
    "TimeXer",
    "iTransformer",
    "DLinear",
    "PatchTST",
    "TimesNet",
    "Autoformer",
]

ZERO_SHOT_MODELS = ["TiRex-2", "Chronos-2", "Toto-2.0", "Moirai"]

CLASSIFICATION_MODELS = [
    "MambaSL",
    "VSFormer",
    "LITE",
    "TimesNet",
    "PatchTST",
    "DLinear",
    "iTransformer",
    "MultiROCKET",
    "MiniROCKET",
    "TabPFN",
]

ANOMALY_MODELS = ["KAN-AD", "Anomaly Trans", "TranAD", "USAD", "OmniAnomaly"]

MODEL_ALIASES = {
    "Chronos-2": ["Chronos-2", "Chronos2"],
    "TiRex-2": ["TiRex-2", "TiRex"],
    "MambaSL": ["MambaSL", "MambaSingleLayer"],
    "KAN-AD": ["KAN-AD", "KANAD"],
    "Anomaly Trans": ["Anomaly Trans", "AnomalyTransformer"],
    "TimeMixer++": ["TimeMixer++", "TimeMixerPP", "TimeMixerPlusPlus"],
}

PENDING_REASON = {
    "OLinear": "PENDING_ADAPT: official code needs Q-matrix and extra encoder layers",
    "TimeMixer++": "PENDING_ADAPT: no stable QAR-ready implementation in current repo",
    "TiRex-2": "MISSING_DEP/WEIGHT: tirex package and pretrained weights unavailable",
    "Chronos-2": "WEIGHT_UNAVAILABLE: server cannot download amazon/chronos-2 from HuggingFace",
    "Toto-2.0": "NOT_IMPLEMENTED: no Toto-2.0 entry in current repo",
    "Moirai": "MISSING_DEP/WEIGHT: uni2ts package and pretrained weights unavailable",
    "MambaSL": "MISSING_DEP: mamba_ssm is not installed on server",
    "VSFormer": "PENDING_ADAPT: not yet connected to QAR compact loader",
    "LITE": "PENDING_ADAPT: not yet connected to QAR compact loader",
    "TabPFN": "PENDING_RUN",
}

FORECAST_LABELS = {
    "predict_2_3": "2->3",
    "predict_4_5": "4->5",
    "predict_5_6": "5->6",
    "predict_8_9": "8->9",
    "phase80": "0->12起点80",
    "default": "forecast",
}


@dataclass(frozen=True)
class MetricSource:
    label: str
    path: Path


def aliases(model: str) -> list[str]:
    values = [model]
    values.extend(MODEL_ALIASES.get(model, []))
    return list(dict.fromkeys(values))


def pending_value(model: str) -> str:
    return PENDING_REASON.get(model, "PENDING_RUN")


def read_csv(path: Path | None) -> pd.DataFrame:
    if path is None or not path.exists():
        return pd.DataFrame()
    return pd.read_csv(path)


def read_many_csv(paths: Iterable[Path]) -> pd.DataFrame:
    frames = []
    for path in paths:
        df = read_csv(path)
        if not df.empty:
            df = df.copy()
            df["_source_csv"] = str(path)
            frames.append(df)
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True, sort=False)


def parse_sources(values: Iterable[str]) -> list[MetricSource]:
    out: list[MetricSource] = []
    for value in values:
        if "=" in value:
            label, raw_path = value.split("=", 1)
        else:
            raw_path = value
            label = Path(raw_path).parent.name
        out.append(MetricSource(label=label.strip(), path=Path(raw_path)))
    return out


def fmt_number(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ""
        try:
            value = float(value)
        except Exception:
            return value
    try:
        number = float(value)
    except Exception:
        return str(value)
    if math.isinf(number):
        return "inf" if number > 0 else "-inf"
    return f"{number:.4f}"


def pick_row(df: pd.DataFrame, dataset: str, model: str) -> pd.Series | None:
    if df.empty or "dataset" not in df.columns or "model" not in df.columns:
        return None
    sub = df[df["dataset"].astype(str) == dataset]
    if sub.empty:
        return None
    for alias in aliases(model):
        hit = sub[sub["model"].astype(str) == alias]
        if not hit.empty:
            if "status" in hit.columns:
                ok = hit[hit["status"].astype(str).isin(["0", "0.0"])]
                if not ok.empty:
                    return ok.iloc[-1]
            return hit.iloc[-1]
    return None


def classification_cell(df: pd.DataFrame, dataset: str, model: str) -> str:
    row = pick_row(df, dataset, model)
    if row is None:
        return pending_value(model)
    return "\n".join(
        [
            f"acc={fmt_number(row.get('accuracy', row.get('acc')))}",
            f"f1={fmt_number(row.get('macro_f1'))}",
            f"TN={fmt_number(row.get('TN'))} TP={fmt_number(row.get('TP'))}",
            f"FN={fmt_number(row.get('FN'))} FP={fmt_number(row.get('FP'))}",
        ]
    )


def forecast_cell(sources: list[MetricSource], dataset: str, model: str) -> str:
    parts: list[str] = []
    for source in sources:
        df = read_csv(source.path)
        row = pick_row(df, dataset, model)
        if row is None:
            continue
        label = FORECAST_LABELS.get(source.label, source.label)
        parts.append(
            f"{label}: mse={fmt_number(row.get('mse'))}, mae={fmt_number(row.get('mae'))}, rmse={fmt_number(row.get('rmse'))}"
        )
    return "\n".join(parts) if parts else pending_value(model)


def anomaly_cell(df: pd.DataFrame, dataset: str, model: str) -> str:
    row = pick_row(df, dataset, model)
    if row is None:
        return pending_value(model)
    return "\n".join(
        [
            f"acc={fmt_number(row.get('accuracy'))} f1={fmt_number(row.get('f1'))}",
            f"P={fmt_number(row.get('precision'))} R={fmt_number(row.get('recall'))}",
            f"TN={fmt_number(row.get('TN'))} TP={fmt_number(row.get('TP'))}",
            f"FN={fmt_number(row.get('FN'))} FP={fmt_number(row.get('FP'))}",
        ]
    )


def write_matrix(
    output: Path,
    classification_csvs: list[Path],
    anomaly_csv: Path | None,
    fullshot_sources: list[MetricSource],
    zeroshot_sources: list[MetricSource],
    title: str,
) -> Path:
    cls_df = read_many_csv(classification_csvs)
    anomaly_df = read_csv(anomaly_csv)

    wb = Workbook()
    ws = wb.active
    ws.title = "benchmark_matrix"

    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_title = PatternFill("solid", fgColor="1F4E78")
    fill_group = PatternFill("solid", fgColor="D9EAF7")
    fill_header = PatternFill("solid", fgColor="E2F0D9")
    fill_pending = PatternFill("solid", fgColor="FFF2CC")
    white = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
    bold = Font(bold=True, name="Microsoft YaHei")
    normal = Font(name="Microsoft YaHei", size=10)

    def set_cell(row: int, col: int, value, fill=None, font=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = border
        cell.font = font or normal
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        if fill:
            cell.fill = fill
        return cell

    row = 1
    set_cell(row, 1, title, fill_title, white)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
    row += 2

    set_cell(row, 1, "数据集", fill_header, bold)
    for col, dataset in enumerate(DATASETS, 2):
        set_cell(row, col, dataset, fill_header, bold)
    row += 1
    set_cell(row, 1, "故障说明", fill_header, bold)
    for col, dataset in enumerate(DATASETS, 2):
        set_cell(row, col, FAULT_DESC.get(dataset, ""), fill_header, bold)
    row += 2

    def write_block(block_title: str, models: list[str], cell_builder):
        nonlocal row
        set_cell(row, 1, block_title, fill_group, bold)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(DATASETS) + 1)
        row += 1
        set_cell(row, 1, "model", fill_header, bold)
        for col, dataset in enumerate(DATASETS, 2):
            set_cell(row, col, dataset, fill_header, bold)
        row += 1
        for model in models:
            set_cell(row, 1, model, None, bold)
            for col, dataset in enumerate(DATASETS, 2):
                value = cell_builder(dataset, model)
                fill = fill_pending if any(
                    value.startswith(prefix)
                    for prefix in ["PENDING", "NOT_IMPLEMENTED", "MISSING_DEP", "WEIGHT_UNAVAILABLE"]
                ) else None
                set_cell(row, col, value, fill=fill)
            row += 1
        row += 2

    write_block(
        "Predictive Maintenance (Full-shot)",
        FULL_SHOT_MODELS,
        lambda dataset, model: forecast_cell(fullshot_sources, dataset, model),
    )
    write_block(
        "Predictive Maintenance (Zero-shot)",
        ZERO_SHOT_MODELS,
        lambda dataset, model: forecast_cell(zeroshot_sources, dataset, model),
    )
    write_block(
        "Fault classification",
        CLASSIFICATION_MODELS,
        lambda dataset, model: classification_cell(cls_df, dataset, model),
    )
    write_block(
        "Anomaly detection",
        ANOMALY_MODELS,
        lambda dataset, model: anomaly_cell(anomaly_df, dataset, model),
    )

    ws.freeze_panes = "B5"
    ws.column_dimensions["A"].width = 24
    for col in range(2, len(DATASETS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 28
    for r in range(1, row + 1):
        ws.row_dimensions[r].height = 58 if r > 4 else 24

    src = wb.create_sheet("source_files")
    src.append(["kind", "label", "path"])
    for csv_path in classification_csvs:
        src.append(["classification", "", str(csv_path)])
    if anomaly_csv:
        src.append(["anomaly", "", str(anomaly_csv)])
    for item in fullshot_sources:
        src.append(["fullshot_forecast", item.label, str(item.path)])
    for item in zeroshot_sources:
        src.append(["zeroshot_forecast", item.label, str(item.path)])
    for row_cells in src.iter_rows():
        for cell in row_cells:
            cell.border = border
            cell.font = normal
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    src.column_dimensions["A"].width = 24
    src.column_dimensions["B"].width = 24
    src.column_dimensions["C"].width = 120

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    return output


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--classification-csv", type=Path, action="append", default=[])
    parser.add_argument("--anomaly-csv", type=Path)
    parser.add_argument(
        "--fullshot-forecast-csv",
        action="append",
        default=[],
        help="Forecast metric CSV. Use label=path to control the label inside cells.",
    )
    parser.add_argument(
        "--zeroshot-forecast-csv",
        action="append",
        default=[],
        help="Zero-shot forecast metric CSV. Use label=path to control the label inside cells.",
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--title", default="QAR benchmark matrix")
    args = parser.parse_args()

    output = write_matrix(
        output=args.output,
        classification_csvs=args.classification_csv,
        anomaly_csv=args.anomaly_csv,
        fullshot_sources=parse_sources(args.fullshot_forecast_csv),
        zeroshot_sources=parse_sources(args.zeroshot_forecast_csv),
        title=args.title,
    )
    print(output)


if __name__ == "__main__":
    main()
